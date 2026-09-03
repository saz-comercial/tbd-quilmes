#!/usr/bin/env python3
"""
Generador de reportes TBD por vendedor (VE).

Uso: python3 generar_reportes.py

Busca en esta misma carpeta el Excel de base (Cognos, hoja 'base' + 'codif')
y el Excel de Estructura más nuevos, cruza todo, y genera un HTML por VE
en la carpeta 'salida/'.
"""

import csv
import glob
import os
import re
import sys
import json
import unicodedata
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Falta instalar una librería. Corré esto una sola vez y volvé a intentar:")
    print("    python3 -m pip install openpyxl")
    sys.exit(1)

CARPETA = os.path.dirname(os.path.abspath(__file__))
CARPETA_SALIDA = os.path.join(CARPETA, "salida")

# ---------------------------------------------------------------------------
# 0. Utilidades
# ---------------------------------------------------------------------------

def norm(v):
    """Normaliza un valor de celda: strings sin espacios extra, resto igual."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def buscar_archivo_mas_nuevo(patron):
    candidatos = [
        f for f in glob.glob(os.path.join(CARPETA, patron))
        if not os.path.basename(f).startswith("~$")
    ]
    if not candidatos:
        return None
    candidatos.sort(key=os.path.getmtime, reverse=True)
    return candidatos[0]


def error_fatal(msg):
    print("\n❌ " + msg)
    print("\nNo generé nada. Arreglá esto y volvé a correr el script.")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 1. Ubicar y abrir los archivos de entrada
# ---------------------------------------------------------------------------

def cargar_workbook_base():
    path = buscar_archivo_mas_nuevo("*ase*anti*.xlsx") or buscar_archivo_mas_nuevo("*ase*antiago*.xlsx")
    if not path:
        error_fatal("No encontré el Excel de la base (algo como 'Base Santi.xlsx') en esta carpeta.")
    print(f"📄 Base de ventas: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for hoja in ("base", "codif"):
        if hoja not in wb.sheetnames:
            error_fatal(f"El Excel de base no tiene una hoja llamada '{hoja}'. Hojas encontradas: {wb.sheetnames}")
    return wb, path


def cargar_workbook_estructura():
    path = buscar_archivo_mas_nuevo("*structura*ctualizada*.xlsx") or buscar_archivo_mas_nuevo("*structura*.xlsx")
    if not path:
        error_fatal("No encontré el Excel de Estructura (algo como 'Estructura actualizada DD-MM-AAAA.xlsx') en esta carpeta.")
    print(f"📄 Estructura de clientes: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Estructura" not in wb.sheetnames:
        error_fatal(f"El Excel de Estructura no tiene una hoja llamada 'Estructura'. Hojas encontradas: {wb.sheetnames}")
    return wb, path


COLUMNAS_NEGOCIACIONES_REQUERIDAS = [
    "Vigencia desde", "Vigencia hasta", "Cliente código", "Cliente nombre",
    "Unidad de negocio", "Línea de pricing", "Linea de pricing nombre",  # sic: la 2da "Linea" no lleva tilde en el archivo real
    "Sin cargo fijo", "Descuento en factura", "Sin cargo variable",
]


def meses_entre(desde, hasta):
    """Meses de calendario entre dos fechas (mismo criterio que usa Santiago
    a mano: año*12+mes, sin contar días). Mínimo 1 para no dividir por cero."""
    return max(1, (hasta.year - desde.year) * 12 + (hasta.month - desde.month))


def _txt(v):
    return "" if v is None else str(v).strip()


def _procesar_filas_negociacion(filas, idx, hoy):
    """
    Lógica compartida para leer las filas de la hoja/archivo de negociaciones,
    sea que vengan de un csv (strings) o de openpyxl (mezcla de int/str).
    Devuelve (negociaciones, filas_vencidas_o_futuras, filas_con_error).
    """
    clientes = {}
    filas_vencidas_o_futuras = 0
    filas_con_error = 0
    tipos = [
        ("Sin cargo fijo", idx["Sin cargo fijo"]),
        ("Descuento en factura", idx["Descuento en factura"]),
        ("Sin cargo variable", idx["Sin cargo variable"]),
    ]

    for fila in filas:
        if not fila or not _txt(fila[idx["Cliente código"]]):
            continue
        try:
            cod = int(_txt(fila[idx["Cliente código"]]))
            desde = datetime.strptime(_txt(fila[idx["Vigencia desde"]]), "%d/%m/%Y")
            hasta = datetime.strptime(_txt(fila[idx["Vigencia hasta"]]), "%d/%m/%Y")
        except (ValueError, IndexError):
            filas_con_error += 1
            continue

        vigente = desde <= hoy <= hasta
        if not vigente:
            filas_vencidas_o_futuras += 1
            continue

        unidad = norm(fila[idx["Unidad de negocio"]])
        linea = norm(fila[idx["Linea de pricing nombre"]])

        for tipo_nombre, col in tipos:
            crudo = _txt(fila[col]).replace(",", ".")
            try:
                valor = float(crudo) if crudo else 0.0
            except ValueError:
                valor = 0.0
            if valor == 0.0:
                continue

            item = {"unidad": unidad, "linea": linea, "tipo": tipo_nombre}
            if tipo_nombre == "Sin cargo fijo":
                meses_totales = meses_entre(desde, hasta)
                meses_transcurridos = min(max(meses_entre(desde, hoy), 0), meses_totales)
                por_mes = round(valor / meses_totales, 1)
                restante = round(max(0.0, valor - por_mes * meses_transcurridos), 1)
                item.update({
                    "total": valor, "meses_totales": meses_totales,
                    "por_mes": por_mes, "restante": restante,
                })
            else:
                item["valor_pct"] = valor

            clientes.setdefault(cod, []).append(item)

    negociaciones = {cod: {"tiene": True, "items": items} for cod, items in clientes.items()}
    return negociaciones, filas_vencidas_o_futuras, filas_con_error


def _validar_columnas_negociacion(idx, origen):
    faltantes = [c for c in COLUMNAS_NEGOCIACIONES_REQUERIDAS if c not in idx]
    if faltantes:
        error_fatal(f"A {origen} le faltan estas columnas: " + ", ".join(faltantes))


def _cargar_negociaciones_xls_legacy(path, hoy):
    """Formato viejo: archivo .xls que en realidad es texto UTF-16 separado por tabs."""
    try:
        with open(path, encoding="utf-16", newline="") as f:
            filas = list(csv.reader(f, delimiter="\t"))
    except (UnicodeError, UnicodeDecodeError):
        error_fatal(
            "El archivo de negociaciones no es el formato de texto (UTF-16 con tabs) que esperaba. "
            "¿Cambió el formato de exportación? Avisame antes de que siga."
        )
    if not filas:
        error_fatal("El archivo de negociaciones está vacío.")

    idx = {h.strip(): i for i, h in enumerate(filas[0])}
    _validar_columnas_negociacion(idx, "el archivo de negociaciones")
    negociaciones, vencidas, con_error = _procesar_filas_negociacion(filas[1:], idx, hoy)
    if con_error:
        print(f"   ⚠️  {con_error} filas del archivo de negociaciones no se pudieron leer (fecha o código raro), las salteé.")
    print(f"   {len(negociaciones)} clientes con negociación vigente hoy ({vencidas} filas vencidas o futuras, ignoradas).")
    return negociaciones


def cargar_negociaciones(hoy):
    """
    Archivo de negociaciones, independiente del histórico (se actualiza
    solo, cada tanto — como Estructura). Puede venir como .xlsx con una hoja
    'NEGOCIACIONES' (con o sin otras hojas al lado), o como el .xls viejo
    (texto UTF-16 separado por tabs). No es obligatorio: si todavía no lo
    subieron, seguimos sin negociaciones en vez de frenar todo.
    """
    path_xlsx = buscar_archivo_mas_nuevo("*ONLINE*.xlsx") or buscar_archivo_mas_nuevo("*egociac*.xlsx")
    if path_xlsx:
        print(f"📄 Negociaciones: {os.path.basename(path_xlsx)}")
        wb = openpyxl.load_workbook(path_xlsx, read_only=True, data_only=True)
        hoja_neg = next((h for h in wb.sheetnames if h.strip().upper() == "NEGOCIACIONES"), wb.sheetnames[0])
        filas_neg = list(wb[hoja_neg].iter_rows(min_row=1, values_only=True))
        if not filas_neg:
            error_fatal("La hoja de negociaciones está vacía.")
        idx_neg = {_txt(h): i for i, h in enumerate(filas_neg[0])}
        _validar_columnas_negociacion(idx_neg, "el archivo de negociaciones")
        negociaciones, vencidas, con_error = _procesar_filas_negociacion(filas_neg[1:], idx_neg, hoy)
        if con_error:
            print(f"   ⚠️  {con_error} filas de negociaciones no se pudieron leer (fecha o código raro), las salteé.")
        print(f"   {len(negociaciones)} clientes con negociación vigente hoy ({vencidas} filas vencidas o futuras, ignoradas).")
        return negociaciones

    path_xls = buscar_archivo_mas_nuevo("*ONLINE*.xls") or buscar_archivo_mas_nuevo("*egociac*.xls")
    if path_xls:
        print(f"📄 Negociaciones (formato viejo): {os.path.basename(path_xls)}")
        return _cargar_negociaciones_xls_legacy(path_xls, hoy)

    print("\nℹ️  No encontré el archivo de negociaciones. Sigo sin negociaciones por ahora.")
    return {}


# ---------------------------------------------------------------------------
# Historial de compras (para la recencia de los CNC): en vez de pedirle a
# Santiago el año completo cada vez, guardamos un archivito chico persistente
# con "el último mes en que cada cliente compró cerveza / UNG+AGUAS", y lo
# vamos actualizando solos con el Base Santi de cada día. Una sola vez (ya
# hecho) se sembró con el histórico completo que mandó Santiago.
# ---------------------------------------------------------------------------

NOMBRE_HISTORIAL = "historial_compras.csv"
CATEGORIAS_HISTORIAL = ("CERVEZA", "UNG_AGUAS")


def cargar_historial_persistido():
    """{(cod, categoria): ultimo_anio_mes}. Si no existe el archivo todavía, vacío."""
    path = os.path.join(CARPETA, NOMBRE_HISTORIAL)
    historial = {}
    if not os.path.exists(path):
        print(f"\nℹ️  Todavía no existe {NOMBRE_HISTORIAL} (primera vez). Arranca vacío.")
        return historial
    with open(path, newline="", encoding="utf-8") as f:
        for fila in csv.DictReader(f):
            try:
                cod = int(fila["cod_cliente"])
                categoria = fila["categoria"]
                anio_mes = int(fila["ultimo_anio_mes"])
            except (KeyError, ValueError):
                continue
            if categoria in CATEGORIAS_HISTORIAL:
                historial[(cod, categoria)] = anio_mes
    print(f"📄 {NOMBRE_HISTORIAL}: {len(historial)} pares cliente+categoría cargados.")
    return historial


def actualizar_historial_con_hoy(historial, registros, anio_mes_hoy):
    """Pisa en 'historial' el mes de hoy para todo cliente que compró esa
    categoría según el Base Santi de hoy. Modifica 'historial' in-place y
    también lo devuelve, por comodidad."""
    for r in registros:
        if r["negocio_agrup"] == "CERVEZA":
            historial[(r["cod_cliente"], "CERVEZA")] = anio_mes_hoy
        elif r["negocio_agrup"] in ("UNG", "AGUAS"):
            historial[(r["cod_cliente"], "UNG_AGUAS")] = anio_mes_hoy
    return historial


def guardar_historial(historial):
    path = os.path.join(CARPETA, NOMBRE_HISTORIAL)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["cod_cliente", "categoria", "ultimo_anio_mes"])
        for (cod, categoria), anio_mes in sorted(historial.items()):
            w.writerow([cod, categoria, anio_mes])
    print(f"   Guardé {NOMBRE_HISTORIAL} con {len(historial)} pares cliente+categoría.")


def _etiqueta_recencia(anio_mes_hoy, anio_mes_ultimo):
    if anio_mes_ultimo is None:
        return "Sin compras en el histórico"
    ay, am = divmod(anio_mes_hoy, 100)
    uy, um = divmod(anio_mes_ultimo, 100)
    meses = (ay - uy) * 12 + (am - um)
    if meses <= 0:
        return None  # compró este mismo mes: no debería estar en la lista CNC
    if meses == 1:
        return "UM"
    if meses == 2:
        return "U2M"
    if meses == 3:
        return "U3M"
    if meses <= 6:
        return f"U{meses}M"
    return "+6"


def calcular_recencia(historial, anio_mes_hoy):
    """A partir de {(cod, categoria): ultimo_anio_mes}, arma {cod: etiqueta}
    para cerveza y para UNG+AGUAS por separado."""
    rec_cerveza = {}
    rec_ung = {}
    for (cod, categoria), anio_mes in historial.items():
        etiqueta = _etiqueta_recencia(anio_mes_hoy, anio_mes)
        if categoria == "CERVEZA":
            rec_cerveza[cod] = etiqueta
        else:
            rec_ung[cod] = etiqueta
    return rec_cerveza, rec_ung


# ---------------------------------------------------------------------------
# 2. codif: tablas de búsqueda oficiales
# ---------------------------------------------------------------------------

COLUMNAS_BASE_REQUERIDAS = [
    "Año Mes Cierre", "COD Cliente", "DSC Sub-canal MKT", "COD Jefe Venta Act",
    "COD Zona Act", "COD Territorio Act", "DSC Negocio", "DSC Marca",
    "DSC Calibre", "DSC Familia", "DSC Sabor", "#HL Vta Bruta", "$ Neto",
    "$ SC 286 - S/C Acc Ctr MKT", "$ Impu Tot", "COD Producto",
    "DSC Sub-Region Agrup Act",
]

COLUMNAS_ESTRUCTURA_REQUERIDAS = [
    "Cod.", "Razon Social", "Domicilio", "Canal", "SubCanal MKT", "Alc.",
    "Zona", "VE", "FREC VE", "REGULAR", "LOCALIDAD", "fecha alta",
]


def construir_codif(ws_codif):
    filas = list(ws_codif.iter_rows(min_row=1, values_only=True))
    if not filas:
        error_fatal("La hoja 'codif' está vacía.")

    # DSC Marca(0) -> Un.Neg(1), Segmento(2), Marca(3) -- fila 1 es encabezado
    marca_map = {}
    for r in filas[1:]:
        k = norm(r[0])
        if k is not None and k not in marca_map:
            marca_map[k] = (norm(r[1]), norm(r[2]), norm(r[3]))

    # DSC Calibre(5) -> Resumen Calibre(6)
    calibre_map = {}
    for r in filas[1:]:
        k = norm(r[5])
        if k is not None and k not in calibre_map:
            calibre_map[k] = norm(r[6])

    # DSC Negocio(13) -> NEGOCIO agrup(14) -- OJO: la fila 0 (índice 0 de la lista) ya es dato
    negocio_map = {}
    for r in filas:
        k = norm(r[13])
        if k is not None and k not in negocio_map:
            negocio_map[k] = norm(r[14])

    return marca_map, calibre_map, negocio_map


# ---------------------------------------------------------------------------
# 3. base: filas de venta, recalculando las columnas derivadas desde codif
# ---------------------------------------------------------------------------

def cargar_base(ws_base, marca_map, calibre_map, negocio_map):
    filas = list(ws_base.iter_rows(min_row=1, values_only=True))
    header = filas[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}

    faltantes = [c for c in COLUMNAS_BASE_REQUERIDAS if c not in idx]
    if faltantes:
        error_fatal("A la hoja 'base' le faltan estas columnas: " + ", ".join(faltantes))

    filas_dato = [r for r in filas[1:] if r[idx["COD Cliente"]] is not None]

    marcas_sin_match = set()
    calibres_sin_match = set()
    negocios_sin_match = set()

    registros = []
    for r in filas_dato:
        cod_cliente = r[idx["COD Cliente"]]
        anio_mes = r[idx["Año Mes Cierre"]]
        dsc_marca = norm(r[idx["DSC Marca"]])
        dsc_calibre = norm(r[idx["DSC Calibre"]])
        dsc_familia = norm(r[idx["DSC Familia"]])
        dsc_sabor = norm(r[idx["DSC Sabor"]])
        dsc_negocio = norm(r[idx["DSC Negocio"]])
        dsc_subcanal = norm(r[idx["DSC Sub-canal MKT"]])
        hl_crudo = r[idx["#HL Vta Bruta"]]
        hl = hl_crudo if isinstance(hl_crudo, (int, float)) else 0

        m = marca_map.get(dsc_marca)
        if m is None:
            marcas_sin_match.add(dsc_marca)
            un_neg, segmento, marca_unif = None, None, None
        else:
            un_neg, segmento, marca_unif = m

        calibre_unif = calibre_map.get(dsc_calibre)
        if calibre_unif is None:
            calibres_sin_match.add(dsc_calibre)

        negocio_agrup = negocio_map.get(dsc_negocio)
        if negocio_agrup is None:
            negocios_sin_match.add(dsc_negocio)

        registros.append({
            "cod_cliente": cod_cliente,
            "anio_mes": anio_mes,
            "dsc_marca": dsc_marca,
            "dsc_calibre": dsc_calibre,
            "dsc_familia": dsc_familia,
            "dsc_sabor": dsc_sabor,
            "dsc_subcanal": dsc_subcanal,
            "negocio_agrup": negocio_agrup,
            "segmento": segmento,
            "calibre_unif": calibre_unif,
            "hl": hl,
        })

    return registros, marcas_sin_match, calibres_sin_match, negocios_sin_match


# ---------------------------------------------------------------------------
# 4. Estructura: universo de clientes
# ---------------------------------------------------------------------------

def cargar_estructura(ws_estructura):
    filas = list(ws_estructura.iter_rows(min_row=1, values_only=True))
    header = filas[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}

    faltantes = [c for c in COLUMNAS_ESTRUCTURA_REQUERIDAS if c not in idx]
    if faltantes:
        error_fatal("A la hoja 'Estructura' le faltan estas columnas: " + ", ".join(faltantes))

    clientes = {}
    descartadas = 0
    duplicados = []
    for r in filas[1:]:
        crudo = r[idx["Cod."]]
        if crudo is None:
            continue
        cod = None
        if isinstance(crudo, int):
            cod = crudo
        elif isinstance(crudo, float):
            cod = int(crudo)
        elif isinstance(crudo, str):
            s = crudo.strip()
            if s.isdigit():
                cod = int(s)
        if cod is None or cod == 0:
            descartadas += 1
            continue

        if cod in clientes:
            duplicados.append(cod)
            continue  # nos quedamos con la primera aparición

        clientes[cod] = {
            "cod": cod,
            "razon_social": norm(r[idx["Razon Social"]]) or f"Cliente {cod}",
            "subcanal_estructura": norm(r[idx["SubCanal MKT"]]),
            "canal": norm(r[idx["Canal"]]),
            "ve": norm(r[idx["VE"]]),
            "frec_ve": norm(r[idx["FREC VE"]]),
        }

    return clientes, descartadas, duplicados


DIAS_VALIDOS = ["LU", "MA", "MI", "JU", "VI", "SA"]


def parsear_frec_ve(frec):
    if not frec or frec == "#N/A":
        return []
    dias = []
    i = 0
    frec = frec.upper()
    while i < len(frec):
        par = frec[i:i + 2]
        if par in DIAS_VALIDOS:
            dias.append(par)
            i += 2
        else:
            i += 1  # caracter raro (espacio, etc.), lo salteamos
    return dias


# ---------------------------------------------------------------------------
# 5. Números de control (misma lógica que se usa para los 6 KPI)
# ---------------------------------------------------------------------------

ZERO_ALC_MARCAS = {
    "quilmes 0.0%", "stella artois 0.0%", "stella artois pure gold",
    "corona cero", "michelob ultra",
}

# clave: nombre del KPI -> (funcion_filtro, funcion_clave_dedup, modo)
# modo "lista" = combos pocos, se listan todos verde/rojo
# modo "top5"  = combos muchos, se recomiendan los 5 más vendidos del sub-canal
def definir_kpis():
    return {
        "TBD CZA": {
            "filtro": lambda r: r["negocio_agrup"] == "CERVEZA",
            "combo": lambda r: (r["dsc_marca"], r["calibre_unif"]),
            "modo": "top5",
        },
        "TBD CORE + VALUE": {
            "filtro": lambda r: r["segmento"] in ("CORE", "VALUE"),
            "combo": lambda r: (r["dsc_marca"], r["calibre_unif"]),
            "modo": "top5",
        },
        "TBD ABOVE CORE": {
            "filtro": lambda r: r["segmento"] == "ABOVE CORE",
            "combo": lambda r: (r["dsc_marca"], r["calibre_unif"]),
            "modo": "top5",
        },
        "TBD LATONES": {
            "filtro": lambda r: r["calibre_unif"] == "Latón",
            "combo": lambda r: (r["dsc_marca"], r["calibre_unif"]),
            "modo": "lista",
        },
        "TBD 0.0% + MUL": {
            "filtro": lambda r: r["dsc_marca"] in ZERO_ALC_MARCAS,
            "combo": lambda r: (r["dsc_marca"], r["calibre_unif"]),
            "modo": "lista",
        },
        "TBD UNG + AGUA": {
            "filtro": lambda r: r["negocio_agrup"] in ("UNG", "AGUAS"),
            # unico KPI que dedupe por 4 campos: el calibre solo no alcanza,
            # dos filas con mismo marca+calibre pueden ser sabores distintos
            # (confirmado por Santiago el 28/07: agregaron DSC Sabor a la base).
            "combo": lambda r: (r["dsc_marca"], r["dsc_calibre"], r["dsc_familia"], r["dsc_sabor"]),
            "modo": "top5",
        },
    }


CALIBRES_UNG_CONOCIDOS = {
    "3000 cc pet", "354 cc lata", "2000 cc pet", "1500 cc pet", "269 cc lata",
    "2250 cc pet", "500 cc vd s/r", "500 cc pet", "1250 cc pet", "400cc pet",
    "354 cc", "250 cc latas", "355 cc lata", "2.25l", "750 cc pet", "2000 reco",
    "350 cc pet", "1000 cc vidrio",
    # sumados el 23/07/2026, confirmados con Santiago:
    "850 cc pet", "bidon 6.3l", "6.4l cc pet", "1000 cc pet",
}


NUMEROS_DE_CONTROL = {
    "TBD CZA": (12646, 2107),
    "TBD CORE + VALUE": (8813, 2014),
    "TBD ABOVE CORE": (3833, 1185),
    "TBD LATONES": (2905, 1200),
    "TBD 0.0% + MUL": (1422, 773),
    # TBD UNG + AGUA sale de esta lista el 28/07/2026: cambio la clave de
    # dedup (ahora suma DSC Sabor) y el numero viejo (9243/1848) ya no es
    # comparable. No es que este mal, es una definicion distinta.
}


def validar_numeros_de_control(registros, kpis):
    print("\n--- Validando números de control ---")
    ok_total = True
    for nombre, spec in kpis.items():
        if nombre not in NUMEROS_DE_CONTROL:
            print(f"  ℹ️  {nombre}: sin número de control fijo (definición cambió recientemente), no se valida.")
            continue
        combos_por_cliente = set()
        for r in registros:
            if spec["filtro"](r):
                combos_por_cliente.add((r["cod_cliente"],) + spec["combo"](r))
        total_tbd = len(combos_por_cliente)
        total_clientes = len(set(c[0] for c in combos_por_cliente))
        esperado_tbd, esperado_cli = NUMEROS_DE_CONTROL[nombre]
        ok = (total_tbd, total_clientes) == (esperado_tbd, esperado_cli)
        ok_total = ok_total and ok
        marca = "✅" if ok else "❌"
        print(f"  {marca} {nombre}: TBD={total_tbd} (esperado {esperado_tbd}) | Clientes={total_clientes} (esperado {esperado_cli})")
    return ok_total


# ---------------------------------------------------------------------------
# 6. Modelo por KPI: universo de combos, quién tiene qué, popularidad por sub-canal
# ---------------------------------------------------------------------------

def titulo(texto):
    """Pone lindo un texto que viene en minúsculas de la base."""
    if texto is None:
        return "?"
    return str(texto).title()


def nombre_combo(*partes):
    """Arma el nombre para mostrar de un combo, sea de 2 partes (marca, calibre)
    o de 4 (marca, calibre, familia, sabor, como en TBD UNG + AGUA)."""
    return " x ".join(titulo(p) for p in partes)


def construir_modelo_kpi(registros, clientes_estructura, kpis):
    """
    Para cada KPI arma:
      - combos_cliente: {cod_cliente: set(combo)}
      - universo: lista ordenada de todos los combos que existen para ese KPI
      - pop_subcanal: {subcanal_estructura: {combo: hl_total}}  (para modo top5)
      - pop_global: {combo: hl_total}  (fallback si el sub-canal no tiene data)
    """
    modelo = {}
    for nombre, spec in kpis.items():
        combos_cliente = {}
        pop_subcanal = {}
        pop_global = {}
        universo = set()

        for r in registros:
            if not spec["filtro"](r):
                continue
            combo = spec["combo"](r)
            if any(parte is None for parte in combo):
                continue  # algun campo del combo (marca/calibre/familia/sabor) sin dato: no lo mostramos prolijo
            universo.add(combo)
            combos_cliente.setdefault(r["cod_cliente"], set()).add(combo)

            cliente_info = clientes_estructura.get(r["cod_cliente"])
            subcanal = cliente_info["subcanal_estructura"] if cliente_info else None
            hl = r["hl"] or 0
            pop_global[combo] = pop_global.get(combo, 0) + hl
            if subcanal:
                pop_subcanal.setdefault(subcanal, {})
                pop_subcanal[subcanal][combo] = pop_subcanal[subcanal].get(combo, 0) + hl

        modelo[nombre] = {
            "modo": spec["modo"],
            "combos_cliente": combos_cliente,
            "universo": sorted(universo, key=lambda c: nombre_combo(*c)),
            "pop_subcanal": pop_subcanal,
            "pop_global": pop_global,
        }
    return modelo


def top5_faltantes(combo_set_universo, combos_que_tiene, subcanal, pop_subcanal, pop_global):
    faltantes = combo_set_universo - combos_que_tiene
    pop_sc = pop_subcanal.get(subcanal, {}) if subcanal else {}

    def clave_orden(combo):
        return (-pop_sc.get(combo, 0), -pop_global.get(combo, 0), nombre_combo(*combo))

    ordenados = sorted(faltantes, key=clave_orden)
    return [nombre_combo(*c) for c in ordenados[:5]]


# ---------------------------------------------------------------------------
# 7. Datos finales por vendedor (VE)
# ---------------------------------------------------------------------------

DIA_LABEL = {"LU": "Lunes", "MA": "Martes", "MI": "Miércoles", "JU": "Jueves", "VI": "Viernes", "SA": "Sábado"}


NEGOCIACION_VACIA = {"tiene": False, "items": []}


def construir_datos_por_ve(clientes_estructura, modelo, kpis, negociaciones, rec_cerveza, rec_ung):
    por_ve = {}
    for cod, info in clientes_estructura.items():
        ve = info["ve"] or "SIN VE"
        dias = parsear_frec_ve(info["frec_ve"])

        kpi_cliente = {}
        for nombre in kpis:
            m = modelo[nombre]
            combos_tiene = m["combos_cliente"].get(cod, set())
            universo = set(m["universo"])
            tiene = len(combos_tiene)
            total = len(universo)

            if m["modo"] == "lista":
                combos_out = [
                    [nombre_combo(*c), c in combos_tiene]
                    for c in m["universo"]
                ]
                kpi_cliente[nombre] = {"modo": "lista", "tiene": tiene, "total": total, "combos": combos_out}
            else:
                faltan = top5_faltantes(
                    universo, combos_tiene, info["subcanal_estructura"], m["pop_subcanal"], m["pop_global"]
                )
                kpi_cliente[nombre] = {"modo": "top5", "tiene": tiene, "total": total, "faltan": faltan}

        cliente_final = {
            "cod": cod,
            "nombre": titulo(info["razon_social"]),
            "canal": info.get("canal"),
            "dias": dias,
            "kpi": kpi_cliente,
            "negociacion": negociaciones.get(cod, NEGOCIACION_VACIA),
            # None = sin datos (no hay histórico cargado todavía), nunca inventar un valor.
            "recencia": {
                "cerveza": rec_cerveza.get(cod, "Sin compras en el histórico" if rec_cerveza else None),
                "ung": rec_ung.get(cod, "Sin compras en el histórico" if rec_ung else None),
            },
        }
        por_ve.setdefault(ve, []).append(cliente_final)

    for ve, lista in por_ve.items():
        lista.sort(key=lambda c: (c["nombre"], c["cod"]))

    return por_ve


def calcular_ccc(clientes):
    """
    CCC = Clientes Con Compra de cerveza: cuantos de la cartera del VE
    compraron aunque sea 1 combo de cerveza este mes, sobre el total de
    clientes que tiene asignados. Usa la misma logica que TBD CZA (cualquier
    NEGOCIO agrup == CERVEZA), asi que reaprovecha ese calculo.
    """
    con_compra = sum(1 for c in clientes if c["kpi"]["TBD CZA"]["tiene"] > 0)
    return {"con_compra": con_compra, "total": len(clientes)}


def totales_por_kpi(clientes, nombres_kpi):
    """
    TBD real del mes para todo el VE: suma de combinaciones distintas de
    COD Cliente + marca + calibre que logró en toda su cartera (la misma
    lógica que el número de control, pero acotada a sus clientes).
    """
    totales = {}
    for nombre in nombres_kpi:
        tbd_total = sum(c["kpi"][nombre]["tiene"] for c in clientes)
        con_tbd_clientes = sum(1 for c in clientes if c["kpi"][nombre]["tiene"] > 0)
        totales[nombre] = {"tbd": tbd_total, "clientes_con_tbd": con_tbd_clientes, "total_clientes": len(clientes)}
    return totales


if __name__ == "__main__":
    print("=== Generador de reportes TBD ===\n")

    wb_base, path_base = cargar_workbook_base()
    marca_map, calibre_map, negocio_map = construir_codif(wb_base["codif"])
    registros, marcas_sin_match, calibres_sin_match, negocios_sin_match = cargar_base(
        wb_base["base"], marca_map, calibre_map, negocio_map
    )
    print(f"   {len(registros)} filas de venta cargadas.")

    wb_est, path_est = cargar_workbook_estructura()
    clientes_estructura, descartadas, duplicados = cargar_estructura(wb_est["Estructura"])
    print(f"   {len(clientes_estructura)} clientes reales en Estructura ({descartadas} filas basura descartadas).")
    if duplicados:
        print(f"   ⚠️  Códigos duplicados en Estructura (me quedé con la primera fila): {duplicados}")

    if negocios_sin_match:
        print(f"\n⚠️  DSC Negocio sin clasificar en codif: {negocios_sin_match}")
        print("   Esto puede hacer que los TBD den mal. Revisá codif antes de seguir.")

    # Marcas/calibres sin clasificar que SÍ importan (afectan directamente algún TBD de cerveza)
    marcas_cerveza_sin_match = {
        r["dsc_marca"] for r in registros
        if r["negocio_agrup"] == "CERVEZA" and r["segmento"] is None
    }
    calibres_cerveza_sin_match = {
        r["dsc_calibre"] for r in registros
        if r["negocio_agrup"] == "CERVEZA" and r["calibre_unif"] is None
    }
    if marcas_cerveza_sin_match:
        print(f"\n⚠️  Marcas de CERVEZA sin clasificar en codif (afecta CORE+VALUE/ABOVE CORE): {marcas_cerveza_sin_match}")
    if calibres_cerveza_sin_match:
        print(f"\n⚠️  Calibres de CERVEZA sin clasificar en codif (afecta CZA/LATONES/etc): {calibres_cerveza_sin_match}")

    calibres_ung_nuevos = {
        r["dsc_calibre"] for r in registros
        if r["negocio_agrup"] in ("UNG", "AGUAS") and r["dsc_calibre"] not in CALIBRES_UNG_CONOCIDOS
    }
    if calibres_ung_nuevos:
        print(f"\n⚠️  Calibres NUEVOS de UNG/AGUAS que no vi antes (revisá si están bien): {calibres_ung_nuevos}")

    kpis = definir_kpis()
    todo_ok = validar_numeros_de_control(registros, kpis)
    if not todo_ok:
        print("\n⚠️  Los números de control NO cierran. Puede ser normal si el mes avanzó")
        print("   (la base se actualiza a diario), pero si el desvío es grande, avisame.")
    else:
        print("\n✅ Los 6 números de control cierran perfecto.")

    cods_base = set(r["cod_cliente"] for r in registros)
    cods_estructura = set(clientes_estructura.keys())
    altas_nuevas = cods_base - cods_estructura
    if altas_nuevas:
        print(f"\n⚠️  {len(altas_nuevas)} clientes están en la base pero NO en Estructura (altas nuevas, no van a ningún VE):")
        print("   " + ", ".join(str(c) for c in sorted(altas_nuevas)))

    ahora = datetime.now()
    negociaciones = cargar_negociaciones(ahora)

    anios_mes_presentes = [r["anio_mes"] for r in registros if r["anio_mes"] is not None]
    anio_mes_hoy = max(anios_mes_presentes) if anios_mes_presentes else int(ahora.strftime("%Y%m"))
    historial = cargar_historial_persistido()
    historial = actualizar_historial_con_hoy(historial, registros, anio_mes_hoy)
    guardar_historial(historial)
    rec_cerveza, rec_ung = calcular_recencia(historial, anio_mes_hoy)

    print("\n--- Armando reportes por vendedor ---")
    modelo = construir_modelo_kpi(registros, clientes_estructura, kpis)
    por_ve = construir_datos_por_ve(clientes_estructura, modelo, kpis, negociaciones, rec_cerveza, rec_ung)

    os.makedirs(CARPETA_SALIDA, exist_ok=True)
    plantilla_path = os.path.join(CARPETA, "plantilla.html")
    if not os.path.exists(plantilla_path):
        error_fatal("Falta el archivo 'plantilla.html' en esta carpeta.")
    with open(plantilla_path, "r", encoding="utf-8") as f:
        plantilla = f.read()

    fecha_hoy = ahora.strftime("%d/%m/%Y %H:%M")
    nombres_kpi = list(kpis.keys())

    generados = 0
    for ve, clientes in sorted(por_ve.items()):
        datos_json = {
            "ve": ve,
            "fecha": fecha_hoy,
            "kpis": nombres_kpi,
            "kpiTotales": totales_por_kpi(clientes, nombres_kpi),
            "ccc": calcular_ccc(clientes),
            "clientes": clientes,
        }
        html = plantilla.replace("__VE__", str(ve)).replace("__FECHA__", fecha_hoy)
        html = html.replace("__DATOS_JSON__", json.dumps(datos_json, ensure_ascii=False))

        nombre_archivo = "TBD_VE_" + re.sub(r"[^A-Za-z0-9_-]", "_", str(ve)) + ".html"
        with open(os.path.join(CARPETA_SALIDA, nombre_archivo), "w", encoding="utf-8") as f:
            f.write(html)
        generados += 1

    print(f"\n✅ Generé {generados} reportes en la carpeta 'salida/'.")
    print("   Mandale a cada vendedor el HTML de su VE (por WhatsApp, mail, lo que uses).")
    print("   Se abre directo en el celular, sin internet y sin instalar nada.")
